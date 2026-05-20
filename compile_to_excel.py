import os
import re
from mistune import create_markdown
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

def extract_table_from_md(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    markdown = create_markdown(plugins=['table'])
    html = markdown(content)
    
    soup = BeautifulSoup(html, 'html.parser')
    table = soup.find('table')
    
    if not table:
        return None, None
    
    headers = [th.text.strip() for th in table.find_all('th')]
    rows = []
    for tr in table.find_all('tr')[1:]:
        row_data = [td.text.strip() for td in tr.find_all('td')]
        if row_data:
            rows.append(row_data)
    
    return headers, rows

def create_summary_sheet(wb, headers, rows):
    ws = wb.create_sheet("Summary")
    
    # Identify indices
    cat_idx = -1
    dist_idx = -1
    steps_idx = -1
    points_idx = -1
    for i, h in enumerate(headers):
        if 'Category' in h: cat_idx = i
        if 'Distance' in h: dist_idx = i
        if 'Steps' in h: steps_idx = i
        if 'Points' in h: points_idx = i
    
    tallies = {
        'Run/Jog': {'Distance': 0.0, 'Steps': 0, 'Points': 0},
        'Cycling': {'Distance': 0.0, 'Steps': 0, 'Points': 0},
        'Steps': {'Distance': 0.0, 'Steps': 0, 'Points': 0}
    }
    
    for row in rows:
        cat = row[cat_idx] if cat_idx != -1 else 'Other'
        try:
            dist = float(row[dist_idx].replace(',', '')) if dist_idx != -1 and row[dist_idx] else 0.0
            steps = int(row[steps_idx].replace(',', '')) if steps_idx != -1 and row[steps_idx] else 0
            pts = int(row[points_idx].replace(',', '')) if points_idx != -1 and row[points_idx] else 0
        except:
            dist, steps, pts = 0.0, 0, 0
            
        if cat in tallies:
            tallies[cat]['Distance'] += dist
            tallies[cat]['Steps'] += steps
            tallies[cat]['Points'] += pts
        else:
            if 'Other' not in tallies: tallies['Other'] = {'Distance': 0.0, 'Steps': 0, 'Points': 0}
            tallies['Other']['Distance'] += dist
            tallies['Other']['Steps'] += steps
            tallies['Other']['Points'] += pts

    ws.append(['Category', 'Total Distance (km)', 'Total Steps', 'Total Points'])
    total_dist = 0
    total_steps = 0
    total_pts = 0
    for cat, data in tallies.items():
        ws.append([cat, round(data['Distance'], 2), data['Steps'], data['Points']])
        total_dist += data['Distance']
        total_steps += data['Steps']
        total_pts += data['Points']
    
    ws.append(['---', '---', '---', '---'])
    ws.append(['GRAND TOTAL', round(total_dist, 2), total_steps, total_pts])

def compile_reports(reports_dir, output_file, members_dir):
    all_headers = []
    all_rows = []
    
    if not os.path.exists(members_dir):
        os.makedirs(members_dir)
    
    for root, dirs, files in os.walk(reports_dir):
        if members_dir in root: continue
        for file in files:
            if file.endswith('.md') and 'Week' in file:
                headers, rows = extract_table_from_md(os.path.join(root, file))
                if headers and rows:
                    if not all_headers: all_headers = headers
                    all_rows.extend(rows)
    
    if not all_rows: return

    # Master File
    wb_master = Workbook()
    ws_master = wb_master.active
    ws_master.title = "All Sessions"
    ws_master.append(all_headers)
    for r in all_rows: ws_master.append(r)
    create_summary_sheet(wb_master, all_headers, all_rows)
    wb_master.save(output_file)

    # Member Files
    profile_idx = next((i for i, h in enumerate(all_headers) if 'Profile' in h), -1)
    member_data = {}
    for r in all_rows:
        name = r[profile_idx]
        if name not in member_data: member_data[name] = []
        member_data[name].append(r)
    
    summary_data = []
    for name, rows in member_data.items():
        wb_member = Workbook()
        ws_member = wb_member.active
        ws_member.title = "Sessions"
        ws_member.append(all_headers)
        for r in rows: ws_member.append(r)
        create_summary_sheet(wb_member, all_headers, rows)
        wb_member.save(os.path.join(members_dir, f"{name}_Activity_Report.xlsx"))
        
        # Collect for global summary
        t = {'Run/Jog': 0, 'Cycling': 0, 'Steps': 0, 'TotalSteps': 0, 'TotalPoints': 0}
        cat_idx = next((i for i, h in enumerate(all_headers) if 'Category' in h), -1)
        steps_idx = next((i for i, h in enumerate(all_headers) if 'Steps' in h), -1)
        pts_idx = next((i for i, h in enumerate(all_headers) if 'Points' in h), -1)
        for r in rows:
            cat = r[cat_idx] if cat_idx != -1 else ''
            try:
                steps = int(r[steps_idx].replace(',', '')) if steps_idx != -1 and r[steps_idx] else 0
                pts = int(r[pts_idx].replace(',', '')) if pts_idx != -1 and r[pts_idx] else 0
            except:
                steps, pts = 0, 0
            if cat in t: t[cat] += pts
            t['TotalSteps'] += steps
            t['TotalPoints'] += pts
        summary_data.append([name, t['Run/Jog'], t['Cycling'], t['Steps'], t['TotalSteps'], t['TotalPoints']])

    # Global Summary Report
    wb_sum = Workbook()
    ws_sum = wb_sum.active
    ws_sum.title = "Leaderboard"
    ws_sum.append(['Member', 'Run/Jog Points', 'Cycling Points', 'Steps Points', 'Cumulative Steps', 'Cumulative Points'])
    for entry in sorted(summary_data, key=lambda x: x[5], reverse=True):
        ws_sum.append(entry)
    wb_sum.save("Strides_in_Sync_2026_Tally.xlsx")
    print("All reports and tallies generated successfully.")

if __name__ == "__main__":
    compile_reports("Reports", "Strides_in_Sync_2026_Compilation.xlsx", os.path.join("Reports", "Members"))
